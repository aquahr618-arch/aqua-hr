"""Load the existing AQUA Excel tracker into the database.

Run on your own computer AFTER schema.sql, e.g.:
    set DATABASE_URL=...      (Supabase connection string)
    python scripts/migrate_excel.py "Professional_HR_Employee_Tracker_Template.xlsx"

Add  --dry-run  to preview what would be inserted without touching the DB.
"""
import os
import sys
import datetime

import openpyxl

MASTER_MAP = {
    "Employee Code": "employee_code", "Full Name": "full_name", "Gender": "gender",
    "DOB": "dob", "Mobile": "mobile", "Official Email": "official_email",
    "Personal Email": "personal_email", "Department": "department",
    "Designation": "designation", "Reporting Manager": "reporting_manager",
    "Date of Joining": "date_of_joining", "Employment Type": "employment_type",
    "Employment Status": "employment_status", "Work Location": "work_location",
    "PAN": "pan", "Aadhaar": "aadhaar", "UAN": "uan", "PF No": "pf_no",
    "ESIC No": "esic_no", "Bank Name": "bank_name", "Account No": "account_no",
    "IFSC": "ifsc", "CTC": "ctc", "Notice Period": "notice_period",
    "Emergency Contact": "emergency_contact", "Emergency Phone": "emergency_phone",
    "Qualification": "qualification", "Experience (Years)": "experience_years",
    "Skills": "skills", "Last Appraisal": "last_appraisal",
    "Next Appraisal": "next_appraisal", "Leave Balance": "leave_balance",
    "Remarks": "remarks",
}
TEXT_FIELDS = {"mobile", "account_no", "emergency_phone", "emergency_contact"}
DATE_FIELDS = {"dob", "date_of_joining"}


def norm(name):
    return " ".join(str(name).split()).upper() if name else ""


def clean(field, value):
    if value is None or value == "":
        return None
    if field in DATE_FIELDS and isinstance(value, (datetime.date, datetime.datetime)):
        return value.date() if isinstance(value, datetime.datetime) else value
    if field in TEXT_FIELDS:
        return str(value).strip()
    return value


def rows_of(ws):
    """Yield each data row of a sheet as a {header: value} dict."""
    headers = [c.value for c in ws[1]]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v in (None, "") for v in r):
            continue
        yield dict(zip(headers, r))


def build(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = {ws.title: ws for ws in wb.worksheets}

    employees = {}      # norm_name -> {id?, fields...}
    order = []

    # 1) roster + ids from Leave Tracker
    for row in rows_of(sheets["Leave Tracker"]):
        name = row.get("Employee Name")
        if not name or not norm(name):
            continue
        n = norm(name)
        if n not in employees:
            employees[n] = {"full_name": " ".join(str(name).split()).title()}
            order.append(n)
        eid = row.get("Employee ID")
        if isinstance(eid, (int, float)):
            employees[n]["id"] = int(eid)

    # 2) enrich with Employee Master details
    for row in rows_of(sheets["Employee Master"]):
        name = row.get("Full Name")
        if not name:
            continue
        n = norm(name)
        if n not in employees:
            employees[n] = {"full_name": str(name).strip()}
            order.append(n)
        for header, field in MASTER_MAP.items():
            employees[n][field] = clean(field, row.get(header))

    # 3) assign ids to anyone still missing one
    used = {e["id"] for e in employees.values() if "id" in e}
    nxt = (max(used) + 1) if used else 1
    for n in order:
        if "id" not in employees[n]:
            employees[n]["id"] = nxt
            nxt += 1

    name_to_id = {n: employees[n]["id"] for n in order}

    # 4) child tables
    leave, payroll, documents = [], [], []
    for row in rows_of(sheets["Leave Tracker"]):
        n = norm(row.get("Employee Name"))
        if n in name_to_id:
            leave.append({"employee_id": name_to_id[n], "cl": row.get("CL"),
                          "sl": row.get("SL"), "leave_taken": row.get("Leave Taken"),
                          "balance": row.get("Balance"), "wfh": row.get("WFH")})
    for row in rows_of(sheets["Payroll"]):
        n = norm(row.get("Employee Name"))
        if n in name_to_id:
            payroll.append({"employee_id": name_to_id[n], "gross": row.get("Gross"),
                            "pf": _s(row.get("PF")), "esic": _s(row.get("ESIC")),
                            "tds": _s(row.get("TDS")), "deductions": _s(row.get("Deductions")),
                            "net_salary": row.get("Net Salary"),
                            "payment_date": _s(row.get("Payment Date"))})
    for row in rows_of(sheets["Documents"]):
        n = norm(row.get("Employee Name"))
        if n in name_to_id:
            documents.append({"employee_id": name_to_id[n], "pan": _s(row.get("PAN")),
                              "aadhaar": _s(row.get("Aadhaar")), "photo": _s(row.get("Photo")),
                              "education": _s(row.get("Education")), "experience": _s(row.get("Experience")),
                              "bank_proof": _s(row.get("Bank Proof")),
                              "offer_letter": _s(row.get("Offer Letter")),
                              "appointment_letter": _s(row.get("Appointment Letter"))})

    emp_rows = [employees[n] for n in order]
    return emp_rows, leave, payroll, documents


def _s(v):
    return None if v in (None, "") else str(v).strip()


def insert_all(url, emp_rows, leave, payroll, documents):
    import psycopg2
    conn = psycopg2.connect(url)
    with conn, conn.cursor() as cur:
        cur.execute("select set_config('app.current_user','migration',false)")
        for e in emp_rows:
            cols = list(e.keys())
            cur.execute(
                f"insert into employees ({','.join(cols)}) "
                f"values ({','.join(['%s']*len(cols))}) on conflict (id) do nothing",
                tuple(e[c] for c in cols),
            )
        # fix the identity sequence so new app inserts don't collide
        cur.execute("select setval(pg_get_serial_sequence('employees','id'), "
                    "(select max(id) from employees))")
        for r in leave:
            _ins(cur, "leave_records", r)
        for r in payroll:
            _ins(cur, "payroll", r)
        for r in documents:
            _ins(cur, "documents", r)
    conn.close()


def _ins(cur, table, row):
    cols = list(row.keys())
    cur.execute(
        f"insert into {table} ({','.join(cols)}) values ({','.join(['%s']*len(cols))})",
        tuple(row[c] for c in cols),
    )


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    dry = "--dry-run" in sys.argv
    if not args:
        sys.exit('Usage: python scripts/migrate_excel.py "<file.xlsx>" [--dry-run]')
    path = args[0]

    emp_rows, leave, payroll, documents = build(path)

    print(f"\nEmployees found: {len(emp_rows)}")
    for e in emp_rows:
        print(f"  #{e['id']:>2}  {e['full_name']}")
    print(f"Leave rows: {len(leave)} | Payroll rows: {len(payroll)} | Document rows: {len(documents)}")

    if dry:
        print("\n[dry-run] Nothing was written to the database.")
        return

    url = os.environ.get("DATABASE_URL")
    if not url:
        sys.exit("Set DATABASE_URL before running without --dry-run.")
    insert_all(url, emp_rows, leave, payroll, documents)
    print("\n✅ Migration complete.")


if __name__ == "__main__":
    main()
